import os
import csv
import time
import gc
import math
from tqdm import tqdm
from openpyxl import load_workbook
import pandas as pd
from transformers import BertModel, BertTokenizer
import torch
import torch.nn as nn
import concurrent.futures

dir_path = r'C:\Users\11435\Desktop\clutter\research\data\others\互动\互动'
out_dir = r'C:\Users\11435\Desktop\clutter\research\data\others\互动\res'

# 加载预训练的BERT模型和分词器
model_path = r'D:\tool\toolkit\nlp\bert-base-chinese'
model = BertModel.from_pretrained(model_path)
tokenizer = BertTokenizer.from_pretrained(model_path)

# 将模型移动到GPU（如果可用）
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
model.to(device)


# 定义函数计算文本嵌入
def get_sentence_embedding(sentence, tokenizer, model):
    inputs = tokenizer(sentence, return_tensors='pt', padding=True, truncation=True, max_length=128)

    # 将输入数据移动到与模型相同的设备（CPU或GPU）
    inputs = {key: val.to(device) for key, val in inputs.items()}

    with torch.no_grad():
        outputs = model(**inputs)
    # 使用CLS token的隐藏状态作为句子的嵌入
    sentence_embedding = outputs.last_hidden_state[:, 0, :]
    return sentence_embedding.squeeze()  # 移除批次维度


def get_cosine_sim(row):
    ques_ebd = get_sentence_embedding(row['Qsubj'].tolist(), tokenizer, model)
    ans_ebd = get_sentence_embedding(row['Reply'].tolist(), tokenizer, model)
    cos = cosine_sim(ques_ebd, ans_ebd)
    return cos.tolist()


def read_compute_write(file_path):
    # 方法效率过低
    df = pd.read_excel(file_path)
    # 批量计算每对问题和回答之间的相似度
    cosine_sim = nn.CosineSimilarity(dim=0, eps=1e-6)
    # cos.item()要修改
    df['similarity'] = df.apply(get_cosine_sim, axis=1)
    df['similarity'][0] = '余弦相似度'
    df[['similarity']].to_excel(os.path.join(out_dir, 'ans_' + os.path.basename(file_path)), index=False)


batch_size = 1000  # 每批处理的行数
cosine_sim = nn.CosineSimilarity(dim=1, eps=1e-6)
needed_header = []
for file_name in os.listdir(dir_path):
    file_path = os.path.join(dir_path, file_name)
    if not os.path.isdir(file_path):
        # 加载工作簿
        workbook = load_workbook(filename=file_path, read_only=True)
        sheet = workbook.active

        # 读取标题行
        headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        header_indices = {header: idx for idx, header in enumerate(headers) if header in needed_header}
        # 初始化一个空的DataFrame来存储所有批次的数据
        all_data = pd.DataFrame(columns=['res'])

        # 分批读取数据
        pbar = tqdm(total=math.ceil(sheet.max_row / batch_size), desc=file_name)
        for row_start in range(2, sheet.max_row + 1, batch_size):
            row_end = min(row_start + batch_size - 1, sheet.max_row)
            data = sheet.iter_rows(min_row=row_start, max_row=row_end, values_only=True)
            batch_df = pd.DataFrame(data, columns=headers)
            batch_df['Qsubj'] = batch_df['Qsubj'].apply(lambda x: x if isinstance(x, str) else '')
            batch_df['Reply'] = batch_df['Reply'].apply(lambda x: x if isinstance(x, str) else '')
            res = res_series = pd.DataFrame(get_cosine_sim(batch_df), columns=['res'])

            all_data = pd.concat([all_data, res], ignore_index=True)

            # 善后工作
            gc.collect()
            torch.cuda.empty_cache()
            pbar.update(1)

            if row_start % 50000 == 0:
                time.sleep(60 * 3)

        pbar.close()
        workbook.close()
        all_data['res'][0] = '余弦相似度'
        all_data.to_excel(os.path.join(out_dir, 'ans_' + os.path.basename(file_path)), index=False)
